import openpyxl

a = openpyxl.load_workbook('../Excel1.xlsx')

print(a.sheetnames)
print("Active Sheet is "+ a.active.title)

# Create Object of any sheet on which you want to work on

result = a['Sheet1']
print(result.title)

#Read data from Excel
print(result['A10'].value)

c1 = result.cell(1,1)
print(c1.value)

c2 = result.cell(column =2, row =9)
print(c2.value)

# Finding Rows having data

rows = result.max_row
column = result.max_column
print("Total Rows are: "+ str(rows))
print("Total Column are: "+ str(column))

for i in range(1, rows+1):
  for j in range(1, column+1):
    c = result.cell(i,j)
    print(c.value)





